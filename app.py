from configparser import ConfigParser
import streamlit as st
import pandas as pd
import os
import re
from datetime import datetime,timedelta
from st_aggrid import AgGrid, GridOptionsBuilder, DataReturnMode, GridUpdateMode
from io import BytesIO
import uuid  # Asegúrate de importar uuid al inicio del archivo
import requests
from bs4 import BeautifulSoup
import msal
from urllib.parse import urlencode, urlparse, parse_qs
import openpyxl
from openpyxl import load_workbook
import altair as alt
import psycopg2
from psycopg2 import sql
from streamlit_option_menu import option_menu
from psycopg2.extras import RealDictCursor
from sqlmodel import create_engine, Session
from sqlmodel import Field, SQLModel
from database import guardar_contenedor_bd
from database import get_session, OCMaui
from sqlmodel import select
from openpyxl.utils.dataframe import dataframe_to_rows


###############################################################################
# 1. CONFIGURACIÓN INICIAL STREAMLIT
###############################################################################
# Título: "AppWebSGA"
# Favicon: URL de la imagen solicitada
st.set_page_config(
    page_title="AppWebSGA",
    page_icon="https://blogger.googleusercontent.com/img/a/AVvXsEgqcaKJ1VLBjTRUn-Jz8DNxGx2xuonGQitE2rZjDm_y_uLKe1_6oi5qMiinWMB91JLtS5IvR4Tj-RU08GEfx7h8FdXAEI5HuNoV9YumyfwyXL5qFQ6MJmZw2sKWqR6LWWT8OuEGEkIRRnS2cqP86TgHOoBVkqPPSIRgnHGa4uSEu4O4gM0iNBb7a8Dunfw1",
    layout="wide"
)

# Configuración de la aplicación (usando valores desde st.secrets)
CLIENT_ID = st.secrets["ms_graph"]["client_id"]
CLIENT_SECRET = st.secrets["ms_graph"]["client_secret"]
TENANT_ID = st.secrets["ms_graph"]["tenant_id"]
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
SCOPES = ["Files.ReadWrite", "User.Read"] # Permisos para leer y escribir en OneDrive


###############################################################################
# 2. CSS PARA PERSONALIZAR LA BARRA LATERAL Y ESTILOS GENERALES
###############################################################################
st.markdown("""
<style>
/* Mover la barra lateral a la derecha */
.css-18e3th9 {
    flex-direction: row-reverse;
}
/* Ajustar ancho (un poco más angosta) */
section[data-testid="stSidebar"] {
    min-width: 280px !important;
    max-width: 280px !important;
    background-color: #001F3F !important;
    color: white !important;
}
/* Forzar color blanco en la sidebar (incluyendo emojis) */
section[data-testid="stSidebar"] * {
    color: #ffffff !important;
}

/* Ajuste general para no superponer */
.css-1laz8t5 {
    flex: 1 1 0%;
}

/* Contenedor principal */
.main-container {
  background-color: #F3F3F3;
  padding: 20px 30px;
  border-radius: 10px;
}

/* Tarjetas (subdivisiones) */
.card {
  background: #fff;
  border: 1px solid #ddd;
  border-radius: 8px;
  padding:16px;
  margin-bottom:16px;
  box-shadow:0 2px 5px rgba(0,0,0,.1);
}

/* Títulos */
h2.title {
  font-size:26px;
  color:#333;
  margin-bottom: 0;
}
p.credit {
  margin: 0;
  font-size: 14px;
  color: #888;
}
p.desc {
  margin-top: 8px;
  font-size:15px;
  color:#444;
}
</style>
""", unsafe_allow_html=True)
# Función para obtener el correo electrónico del usuario desde Microsoft Graph
def get_user_email(access_token):
    url = "https://graph.microsoft.com/v1.0/me"
    headers = {"Authorization": f"Bearer {access_token}"}
    response = requests.get(url, headers=headers)
    data = response.json()
    if "mail" in data:
        return data["mail"]
    elif "userPrincipalName" in data:  # Si "mail" no está disponible, se utiliza el nombre principal del usuario
        return data["userPrincipalName"]
    return None


# Función para listar archivos de OneDrive
def list_onedrive_files(access_token):
    """Obtiene la lista de archivos desde OneDrive usando el token de acceso."""
    url = "https://graph.microsoft.com/v1.0/me/drive/root/children"
    headers = {"Authorization": f"Bearer {access_token}"}
    response = requests.get(url, headers=headers)
    data = response.json()
    if "value" in data:
        return data["value"]  # Lista de archivos
    return None

redirect_uri = "https://localhost:8501/signout-oidc/"
# redirect_uri = st.secrets["ms_graph"]["redirect_uri"]

# Función para generar el enlace de autorización con la redirección correcta
def get_authorization_url():
    app = msal.PublicClientApplication(CLIENT_ID, authority=AUTHORITY)
    auth_url = app.get_authorization_request_url(SCOPES, redirect_uri=redirect_uri)
    print(f"auth_url$:${auth_url}")
    return auth_url
# Función para intercambiar el código de autorización por un token de acceso
def get_access_token_from_code(auth_code):
    print(f"${auth_code}")
    app = msal.PublicClientApplication(CLIENT_ID, authority=AUTHORITY)
    result = app.acquire_token_by_authorization_code(auth_code, scopes=SCOPES, redirect_uri=redirect_uri)
    print(f"${result}")
    # Verificar si se obtuvo el token
    if result is None:
        st.error("El resultado de la autenticación es None. Verifica los parámetros de la solicitud.")
        return None
    
    if "access_token" in result:
        return result["access_token"]
    else:
        st.error(f"Error al obtener el token de acceso: {result.get('error_description')}")
        return None

def login_button():
    auth_url = get_authorization_url()
    

# Función para obtener el código de autorización desde la URL
def get_auth_code_from_url():
    url = st.query_params  # Obtener los parámetros de la URL
    if 'code' in url:
        return url['code'][0]  # Extraer el código de autorización
    return None

###############################################################################
# 3. TABLA INTERACTIVA SIN AUTO-ACTUALIZACIÓN (NO_UPDATE)
###############################################################################

# Función para mostrar y exportar el DataFrame
def interactive_table_no_autoupdate(df: pd.DataFrame, key: str = None) -> pd.DataFrame:
    """
    Muestra un DataFrame con st_aggrid usando update_mode=NO_UPDATE:
      - La tabla es interactiva: los usuarios pueden filtrar y ordenar los datos
      - Exportación a Excel para descargar los datos.
    """
    # Configurar la tabla interactiva con AgGrid
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(filter=True)  # Habilitar filtros sin permitir edición
    gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=20)
    gb.configure_grid_options(
        paginationPageSize=20,
        paginationPageSizeOptions=[20, 50, 100, 200, 500]
    )
    gb.configure_side_bar()

    grid_options = gb.build()
    grid_response = AgGrid(
        df,
        gridOptions=grid_options,
        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
        update_mode=GridUpdateMode.MODEL_CHANGED,
        theme="blue",
        key=key
    )

    # Botón para exportar la tabla a Excel
    if st.button("Exportar a Excel", key=f"exportar_excel_{key}"):
        # Crear el archivo Excel en memoria (sin guardarlo en el disco)
        output = BytesIO()

        # Usamos el motor 'openpyxl' para crear el archivo Excel
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Datos")
            # No es necesario llamar a writer.save(), ya que openpyxl lo maneja automáticamente

        output.seek(0)  # Volver al inicio del archivo

        # Crear el botón de descarga usando `st.download_button`
        st.download_button(
            label="Descargar archivo Excel",
            data=output,
            file_name=f"{key}_editado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return df


###############################################################################
# 4. FUNCIONES AUXILIARES DE NEGOCIO
###############################################################################
def set_directories():
    """Crea la carpeta local 'DATA_MAUI_PJLT' si no existe."""
    os.makedirs('DATA_MAUI_PJLT', exist_ok=True)

def extraer_descripcion(descripcion):
    valores_bano = [
        "T. Baño Entero", "T. Baño Stretch", "T. Baño Corto",
        "T. Baño Microfibra", "Traje de Baño", "T.BAÑO"
    ]
    if isinstance(descripcion, str):
        if descripcion[:2] == " *":
            match = re.search(r'Pack\s([A-Za-z\s]+)\s\d', descripcion)
            resultado = match.group(1).strip() if match else None
        else:
            match = re.search(r'\*\s([A-Za-z\s]+)\s\d', descripcion)
            resultado = match.group(1).strip() if match else None

        if resultado in valores_bano:
            return "Traje de Baño"
        return resultado
    else:
        return None

def extraer_codigo_marca(descripcion, subfamilia):
    if not isinstance(descripcion, str) or not isinstance(subfamilia, str):
        return None
    patron = rf"{re.escape(subfamilia)}\s+([\w\-]+)"
    match = re.search(patron, descripcion, flags=re.IGNORECASE)
    if match:
        code = match.group(1)
        if "-" in code:
            code = code.split("-")[0]
        return code.strip()
    return None

def calcular_marca(codigo_marca):
    if not isinstance(codigo_marca, str) or not codigo_marca.strip():
        return None
    match = re.search(r'(\d)', codigo_marca)
    if match:
        d = match.group(1)
        if d == '5':
            return "Maui"
        elif d == '6':
            return "Rip Curl"
        elif d in ['4', '7']:
            return "Rusty y Otros"
    return None

def calcular_zona(marca):
    if marca == "Maui":
        return "B2.ME02 - B2.ME03 - B2.ME04.C09 a B2.ME04.C15"
    elif marca == "Rip Curl":
        return "B1.ME03 - B1.M0E04"
    elif marca == "Rusty y Otros":
        return "B2.ME04.C01 a B2.ME04.C05"
    return None
#Función auxiliar para calcular Tipo_Pack
def calcular_tipo_pack(descripcion: str) -> str:
    """
    Retorna "Inner" si los dos primeros caracteres son " *",
    de lo contrario retorna "Unidad".
    """
    if isinstance(descripcion, str) and len(descripcion) >= 2 and descripcion[:2] == " *":
        return "Inner"
    return "Unidad"
def calcular_factor_por_caja(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calcula la columna 'Factor por Caja' en el dataframe `df` basado en:
    - Agrupación por la columna 'Prtnum Padre'
    - Suma de la columna 'Cantidad Empleada'
    """
    # Verificamos que existan las columnas requeridas
    required_columns = ["Prtnum Padre", "Cantidad Empleada"]
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Falta la columna obligatoria '{col}' en df_curva_articulo")

    # Usamos 'transform' para sumar en cada grupo el valor de 'Cantidad Empleada'
    df["Factor por Caja"] = df.groupby("Prtnum Padre")["Cantidad Empleada"].transform("sum")

    return df
def calcular_factor_caja(df_consolidado: pd.DataFrame, 
                         df_curva_articulo: pd.DataFrame) -> pd.DataFrame:
    """
    1. Toma el valor de 'Num Producto' en df_consolidado,
    2. Busca coincidencia en df_curva_articulo['Prtnum Padre']
    3. Retorna el 'Factor por Caja' si hay coincidencia; si no, 1.
    """
    # Verificar que existan las columnas necesarias
    if "Num Producto" not in df_consolidado.columns:
        raise ValueError("No se encontró la columna 'Num Producto' en df_consolidado.")
    if "Prtnum Padre" not in df_curva_articulo.columns:
        raise ValueError("No se encontró la columna 'Prtnum Padre' en df_curva_articulo.")
    if "Factor por Caja" not in df_curva_articulo.columns:
        raise ValueError("No se encontró la columna 'Factor por Caja' en df_curva_articulo.")

    # Creamos un diccionario para mapear: { Prtnum Padre -> Factor por Caja }
    mapping = dict(zip(df_curva_articulo["Prtnum Padre"], df_curva_articulo["Factor por Caja"]))

    # Para cada fila de df_consolidado, busca 'Num Producto' en mapping; si no existe, asigna 1
    df_consolidado["Factor_Caja"] = df_consolidado["Num Producto"].map(mapping).fillna(1)
    return df_consolidado


def calcular_qty_inners(df_consolidado: pd.DataFrame) -> pd.DataFrame:
    """
    1. Si Tipo_Pack == "Inner" => Qty_Inners = Cantidad Esperada
    2. Si Tipo_Pack == "Unidad" => Qty_Inners = Cant Cajas
    """
    # Verificar que existan las columnas necesarias
    required_cols = ["Tipo_Pack", "Cantidad Esperada", "Cant Cajas"]
    for col in required_cols:
        if col not in df_consolidado.columns:
            raise ValueError(f"No se encontró la columna '{col}' en df_consolidado.")

    # Aplicar la lógica
    df_consolidado["Qty_Inners"] = df_consolidado.apply(
        lambda row: row["Cantidad Esperada"] if row["Tipo_Pack"] == "Inner" 
                    else row["Cant Cajas"],
        axis=1
    )
    return df_consolidado


def calcular_qty_unidades(df_consolidado: pd.DataFrame) -> pd.DataFrame:
    """
    1. Si Tipo_Pack == "Inner" => Qty_Unidades = Qty_Inners * Factor_Caja
    2. Si Tipo_Pack == "Unidad" => Qty_Unidades = Cantidad Esperada
    """
    # Verificar que existan las columnas necesarias
    required_cols = ["Tipo_Pack", "Factor_Caja", "Qty_Inners", "Cantidad Esperada"]
    for col in required_cols:
        if col not in df_consolidado.columns:
            raise ValueError(f"No se encontró la columna '{col}' en df_consolidado.")

    # Aplicar la lógica
    df_consolidado["Qty_Unidades"] = df_consolidado.apply(
        lambda row: row["Qty_Inners"] * row["Factor_Caja"] if row["Tipo_Pack"] == "Inner"
                    else row["Cantidad Esperada"],
        axis=1
    )
    return df_consolidado
def generar_df_f_expl_unid(df_consolidado: pd.DataFrame) -> pd.DataFrame:
    """
    Genera el DataFrame df_f_expl_unid tomando solo las filas con 'Tipo_Pack' == 'Unidad'
    y ajustando las columnas al orden y nombre requeridos.
    """
    # Filtrar los registros con Tipo_Pack == 'Unidad'
    df_unid = df_consolidado[df_consolidado["Tipo_Pack"] == "Unidad"].copy()

    # Verificar columnas base
    columnas_requeridas = [
        "No Factura",
        "Num Producto",
        "Descripcion",          # o "Descripción", según tu CSV
        "Qty_Unidades",
        "Familia De Producto",
        "Zona"
    ]
    for col in columnas_requeridas:
        if col not in df_unid.columns:
            raise ValueError(f"Falta la columna '{col}' en df_consolidado para generar df_f_expl_unid.")

    # Crear nueva columna Observaciones (vacía)
    df_unid["Observaciones"] = ""

    # Renombrar y reordenar columnas
    df_unid.rename(
        columns={
            "No Factura": "No Factura",
            "Num Producto": "Código Hijo",
            "Descripcion": "Descripción",
            "Qty_Unidades": "Cantidad Und",
            "Familia De Producto": "Familia",
            "Zona": "Piso",
        },
        inplace=True
    )

    # Orden final
    df_unid = df_unid[
        ["No Factura",
         "Código Hijo",
         "Descripción",
         "Cantidad Und",
         "Familia",
         "Observaciones",
         "Piso"]
    ]

    return df_unid

def mostrar_y_descargar_dataframe(df: pd.DataFrame, nombre: str):
    # Mostrar con AgGrid (o st.dataframe)
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_pagination(paginationAutoPageSize=True)
    gb.configure_side_bar()
    grid_options = gb.build()

    AgGrid(
        df,
        gridOptions=grid_options,
        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
        update_mode=GridUpdateMode.NO_UPDATE,
        theme="blue"
    )

    # Botón para exportar a Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nombre)
    output.seek(0)

    st.download_button(
        label="Descargar en Excel",
        data=output,
        file_name=f"{nombre}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# ------------------------------------------------------------------------------
# 2. Generar DataFrames de Salida (ejemplos)
# ------------------------------------------------------------------------------
def generar_df_f_expl_unid(df_consolidado: pd.DataFrame) -> pd.DataFrame:
    """
    Filtra Tipo_Pack == 'Unidad' y reordena/renombra columnas:
      1) No Factura -> (misma)
      2) Código Hijo -> Num Producto
      3) Descripción -> Descripcion
      4) Cantidad Und -> Qty_Unidades
      5) Familia -> Familia De Producto
      6) Observaciones -> (columna vacía)
      7) Piso -> Zona
    """
    df_unid = df_consolidado[df_consolidado["Tipo_Pack"] == "Unidad"].copy()
    required_cols = [
        "No Factura",
        "Num Producto",
        "Descripcion",
        "Qty_Unidades",
        "Familia De Producto",
        "Zona"
    ]
    for col in required_cols:
        if col not in df_unid.columns:
            raise ValueError(f"Falta la columna '{col}' en df_consolidado.")

    df_unid["Observaciones"] = ""  # Columna vacía
    df_unid.rename(columns={
        "Num Producto": "Código Hijo",
        "Descripcion": "Descripción",
        "Qty_Unidades": "Cantidad Und",
        "Familia De Producto": "Familia",
        "Zona": "Piso"
    }, inplace=True)

    df_unid = df_unid[[
        "No Factura",
        "Código Hijo",
        "Descripción",
        "Cantidad Und",
        "Familia",
        "Observaciones",
        "Piso"
    ]]
    return df_unid

def generar_df_f_recepcion(df_consolidado: pd.DataFrame) -> pd.DataFrame:
    """
    Genera el DataFrame df_f_recepcion con columnas específicas
    basadas en df_consolidado, solo para los registros cuyo valor en 
    la columna 'Tipo_Pack' sea 'Inner'. También añade la columna 'Piso',
    que toma el valor de la columna 'Zona' del df_consolidado.
    """

    # Filtrar los registros donde "Tipo_Pack" sea "Inner"
    df_consolidado_inner = df_consolidado[df_consolidado["Tipo_Pack"] == "Inner"]

    # Crear un DataFrame nuevo con las columnas indicadas
    df_recepcion = pd.DataFrame()

    # 1) "No Factura"
    df_recepcion["No Factura"] = df_consolidado_inner["No Factura"]

    # 2) "Familia" (viene de "Familia De Producto")
    df_recepcion["Familia"] = df_consolidado_inner["Familia De Producto"]

    # 3) "Código Padre" (viene de "Num Producto")
    df_recepcion["Código Padre"] = df_consolidado_inner["Num Producto"]

    # 4) "Descripcion" (viene de "Descripcion")
    df_recepcion["Descripcion"] = df_consolidado_inner["Descripcion"]

    # 5) "Cant. Inner" (viene de "Qty_Inners")
    df_recepcion["Cant. Inner"] = df_consolidado_inner["Qty_Inners"]

    # 6) "Cant. Unidades" (viene de "Qty_Unidades")
    df_recepcion["Cant. Unidades"] = df_consolidado_inner["Qty_Unidades"]

    # 7) "Cant. Alm. Inner (20%)" => entero de (Qty_Inners * 0.2)
    df_recepcion["Cant. Alm. Inner (20%)"] = (
        df_consolidado_inner["Qty_Inners"] * 0.2
    ).astype(int)

    # 8) "Cant. Alm. Und (80%)" => entero de (Qty_Inners * 0.8)
    df_recepcion["Cant. Alm. Und (80%)"] = (
        df_consolidado_inner["Qty_Inners"] * 0.8
    ).astype(int)

    # 9) "Sub Familia" (viene de "Subfamilias")
    df_recepcion["Sub Familia"] = df_consolidado_inner["Subfamilias"]

    # 10) Añadir la columna "Piso" con el mismo valor que la columna "Zona"
    df_recepcion["Piso"] = df_consolidado_inner["Zona"]

    # Eliminar las filas donde "Tipo_Pack" sea "Unidad" para evitar filas vacías
    df_recepcion = df_recepcion[df_consolidado["Tipo_Pack"] != "Unidad"]

    return df_recepcion


def generar_df_expl_inner(df_f_recepcion: pd.DataFrame, 
                          df_curva_articulo: pd.DataFrame) -> pd.DataFrame:
    """
    Genera el DataFrame df_expl_inner uniendo df_f_recepcion con df_curva_articulo.
    - 'Código Padre' en df_f_recepción se asume que coincide con 'Prtnum Padre' en df_curva_articulo
    - Calcula las columnas según la estructura deseada.
    """

    # Unir por "Código Padre" (izquierda) con "Prtnum Padre" (derecha)
    df_merged = df_f_recepcion.merge(
        df_curva_articulo, 
        left_on="Código Padre", 
        right_on="Prtnum Padre", 
        how="left"
    )

    # Crear nuevo df con las columnas en el orden deseado:
    df_expl_inner = pd.DataFrame()

    df_expl_inner["No Factura"] = df_merged["No Factura"]
    df_expl_inner["Código Padre"] = df_merged["Código Padre"]

    # "Can. Cód. Padre (Inner)" = 'Cant. Alm. Und (80%)' de df_f_recepción
    df_expl_inner["Can. Cód. Padre (Inner)"] = df_merged["Cant. Alm. Und (80%)"]

    # "Código Hijo" = 'Prtnum Hijo' de la curva
    df_expl_inner["Código Hijo"] = df_merged["Prtnum Hijo"]

    # "Factor Caja" = 'Factor' de la curva (asegúrate del nombre de tu columna)
    df_expl_inner["Factor Caja"] = df_merged["Factor por Caja"]

    # "Factor Hijo (Und)" = 'Cantidad Empleada' de la curva
    df_expl_inner["Factor Hijo (Und)"] = df_merged["Cantidad Empleada"]

    # "Cant. Alm. Und" = producto de "Factor Hijo (Und)" × "Can. Cód. Padre (Inner)"
    df_expl_inner["Cant. Alm. Und"] = (
        df_expl_inner["Factor Hijo (Und)"] 
        * df_expl_inner["Can. Cód. Padre (Inner)"]
    )

    # "Cant. Fisico. Und" => columna en blanco
    df_expl_inner["Cant. Fisico. Und"] = ""

    # "Piso" => viene de df_f_recepción
    df_expl_inner["Piso"] = df_merged["Piso"]

    return df_expl_inner


def parsear_fecha(valor):
    """
    Convierte un número de fecha de Excel (por ejemplo, 45505.042083333334) en formato datetime.
    Si no puede convertir el valor, lo devuelve como NaT.
    """
    try:
        if pd.isna(valor):
            return None
        if isinstance(valor, (float, int)):  # Si es un valor numérico de Excel
            base = datetime(1899, 12, 30)  # Base de fecha de Excel
            delta = timedelta(days=valor)  # Días desde la fecha base
            return base + delta  # Sumar la diferencia para obtener la fecha
        else:
            # Si es una fecha en formato de texto, intentamos convertirla a datetime
            return pd.to_datetime(valor, errors='coerce')  # Coerce convierte errores en NaT
    except Exception as e:
        return None
    

def mostrar_resumen_oc(df_consolidado):
    # 1. Cálculos
    total_inners = df_consolidado["Qty_Inners"].sum()
    total_unidades = df_consolidado["Qty_Unidades"].sum()

    df_totales_marca = (
        df_consolidado
        .groupby("Marca", as_index=False)[["Qty_Inners", "Qty_Unidades"]]
        .sum()
    )
    df_totales_zona = (
        df_consolidado
        .groupby("Zona", as_index=False)[["Qty_Inners", "Qty_Unidades"]]
        .sum()
    )

    # 2. Mostrar resultados
    st.markdown("#### Resumen Totales")
    col1, col2 = st.columns(2)
    with col1:
        st.metric("Total Inners", f"{total_inners:,.0f}")
    with col2:
        st.metric("Total Unidades", f"{total_unidades:,.0f}")

    # Tablas auxiliares
    st.markdown("#### Totales por Marca")
    st.table(df_totales_marca)  # Puedes usar st.table o st.dataframe

    st.markdown("#### Totales por Zona")
    st.table(df_totales_zona)


###############################################################################
# 5. MENÚ LATERAL (A LA DERECHA) CON ICONOS BLANCOS
###############################################################################
MENU_OPCIONES = {
    "Inicio": "house",
    "Realizar Análisis": "search",
    "Registro de OC´s": "pencil-square",
    "Consultar BD": "folder",
    "Salir": "door-closed"
}


def set_menu_selection():
    """Se asegura de que 'menu_selected' exista en st.session_state."""
    if "menu_selected" not in st.session_state:
        st.session_state["menu_selected"] = "Inicio"

def radio_menu_con_iconos():
    MENU_OPCIONES = {
        "Inicio": "house",
        "Realizar Análisis": "search",
        "Registro de OC´s": "pencil-square",
        "Consultar BD": "folder",
        "Salir": "door-closed"
    }
    names = list(MENU_OPCIONES.keys())   # ["Inicio", "Realizar Análisis", ...]
    icons = list(MENU_OPCIONES.values()) # ["house", "search", ...]

    with st.sidebar:
        st.image(
            "https://blogger.googleusercontent.com/img/a/AVvXsEgG46LCtcs4m21eiV-0iDqPHZpdfuEEQrJAqwKNY2WPZWdaC1eoAokveaOPXpitT2a_vsKB7zCnxhRfadp0Edz0q5CcfERwYVzrTZSIeeay_o31XrYlqRxocgNau6kWPjAA61uD42zK--pQlZ6wsyIp97mKU53kHZO-yZXjp_wMNv6Coo_CMiitELregplf=w320-h320",
            use_container_width=True
        )
        # Menú con option_menu
        seleccion = option_menu(
            menu_title="",
            options=names,
            icons=icons,
            menu_icon="cast",
            orientation="vertical",
            key="menu_selected",  # <--- Aquí es donde dejamos que Streamlit maneje el estado
            styles={
                "container": {"background-color": "#001F3F"},
                "nav-link": {
                    "font-size": "14px",
                    "text-align": "left",
                    "margin": "0px",
                    "--hover-color": "#0066CC",
                    "color": "white",
                },
                "nav-link-selected": {
                    "background-color": "#0078d4",
                    "color": "white"
                },
            }
        )

        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("<p style='text-align:center; font-size:12px; margin-top:30px;'>Developed by: PJLT</p>", 
                    unsafe_allow_html=True)

    return seleccion


###############################################################################
# 6. LECTURA DE token
###############################################################################
def get_access_token():
    """
    Obtiene un token de Microsoft Graph usando Client Credentials,
    leyendo las credenciales desde st.secrets.
    """
    # Lee la sección [ms_graph] definida en secrets.toml
    tenant_id = st.secrets["ms_graph"]["tenant_id"]
    client_id = st.secrets["ms_graph"]["client_id"]
    client_secret = st.secrets["ms_graph"]["client_secret"]

    authority_url = f"https://login.microsoftonline.com/{tenant_id}"
    scopes = ["https://graph.microsoft.com/.default"]

    app = msal.ConfidentialClientApplication(
        client_id=client_id,
        client_credential=client_secret,
        authority=authority_url
    )
    result = app.acquire_token_for_client(scopes=scopes)

    if not result:
        st.error("No se obtuvo ninguna respuesta al solicitar el token.")
    else:
        # Si 'result' no es None, verifica si tiene 'access_token' o si contiene 'error'
        if 'access_token' in result:
            # Maneja tu lógica con el token
            access_token = result['access_token']
            st.success("Token obtenido correctamente.")
        else:
            # Muestra el error completo o un mensaje amigable
            st.error(f"No se obtuvo el token. Respuesta devuelta: {result}")
    return result["access_token"]

###############################################################################
# 6. PÁGINAS / SECCIONES
###############################################################################
def page_home():
    col1, col2 = st.columns([1,3])
    with col1:
        # Logotipo en la sección de inicio
        st.image("https://www.dinet.com.pe/img/logo-dinet.png", width=170)
    with col2:
        st.markdown("<h2 class='title'>Sistema de Gestión de Abastecimiento - MAUI</h2>", unsafe_allow_html=True)
        st.markdown("<p class='credit'>Developed by: <b>PJLT</b></p>", unsafe_allow_html=True)
        st.markdown("<p class='desc'>Este sistema realiza análisis y registro de datos de abastecimiento.</p>", unsafe_allow_html=True)

    st.markdown("---")
    st.info("Selecciona una opción en la barra lateral para comenzar.")


def page_consultar_bd():
    """
    Esta función ahora hace 2 cosas:
      1) Conectar a OneDrive y listar/cargar un archivo (tal como ya lo hacías).
      2) Mostrar un CRUD (Create, Read, Update, Delete) sobre la tabla OCMaui.
    """

    st.markdown("## Conectar a OneDrive con Microsoft Graph (Protegido con st.secrets)")

    # -------------------------------------------------------------------------
    # SECCIÓN 1: Lógica EXISTENTE para conectar y listar archivos de OneDrive
    # -------------------------------------------------------------------------
    token = get_access_token()
    if not token:
        return  # Error en obtención de token

    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get("https://graph.microsoft.com/v1.0/me/drive/root/children", headers=headers)
    data = resp.json()

    if "value" not in data:
        st.error(f"No se encontraron archivos: {data}")
        return

    archivos = data["value"]
    if not archivos:
        st.warning("No hay archivos en la carpeta raíz de OneDrive.")
    else:
        nombres = [item["name"] for item in archivos]
        seleccionado = st.selectbox("Seleccionar archivo", nombres)

        if st.button("Cargar archivo"):
            # Buscamos el item
            item = next((i for i in archivos if i["name"] == seleccionado), None)
            if not item:
                st.warning("No se encontró el archivo en la respuesta.")
                return

            download_url = item.get("@microsoft.graph.downloadUrl")
            if not download_url:
                st.warning("No hay enlace de descarga.")
                return

            r_file = requests.get(download_url)
            r_file.raise_for_status()

            # Asumimos un Excel .xlsx de ejemplo
            try:
                df = pd.read_excel(BytesIO(r_file.content), engine="openpyxl")
                st.success("Archivo leído con éxito. Vista previa:")
                st.dataframe(df.head(20))
            except Exception as e:
                st.error(f"Error al leer el archivo: {e}")


    # -------------------------------------------------------------------------
    # SECCIÓN 2: CRUD en la tabla oc_maui
    # -------------------------------------------------------------------------
    st.markdown("---")
    st.markdown("## CRUD en la Tabla oc_maui (Ejemplo)")

    # 1) LEER (READ) todos los registros actuales
    with get_session() as session:
        statement = select(OCMaui).order_by(OCMaui.id)
        results = session.exec(statement).all()

    # Convertir los registros en un DataFrame para visualizarlos
    if results:
        df_db = pd.DataFrame([row.dict() for row in results])  # row.dict() si usas SQLModel
    else:
        df_db = pd.DataFrame(columns=["id", "shipment", "referencia", "..."])

    st.subheader("Registros en oc_maui")
    st.dataframe(df_db)  # Mostramos el dataframe con los registros actuales

    # 2) CREAR (CREATE)
    st.subheader("Crear un nuevo registro")
    with st.form("crear_registro_form", clear_on_submit=True):
        # Ajusta los campos que quieres solicitar; 
        # por ejemplo, shipment, referencia, fecha_recepcion, etc.
        new_shipment = st.text_input("Shipment")
        new_referencia = st.text_input("Referencia")
        new_fecha = st.text_input("Fecha Recepción (YYYY-MM-DD)", value=str(datetime.now().date()))
        new_marca = st.text_input("Marca")
        # ... añade más campos según tu modelo ...

        submitted_create = st.form_submit_button("Crear")
        if submitted_create:
            if new_shipment and new_referencia:
                try:
                    with get_session() as session:
                        nuevo_registro = OCMaui(
                            shipment=new_shipment,
                            referencia=new_referencia,
                            fecha_recepcion=new_fecha,
                            marca=new_marca,
                            # ... etc. ...
                        )
                        session.add(nuevo_registro)
                        session.commit()
                        st.success("¡Registro creado exitosamente!")
                    st.experimental_rerun()  # refresca la página para ver el nuevo registro
                except Exception as e:
                    st.error(f"Error al crear registro: {e}")
            else:
                st.warning("Por favor llena al menos Shipment y Referencia.")

    # 3) EDITAR (UPDATE)
    st.subheader("Editar un registro existente (por ID)")
    # a) Pedir el ID a editar
    with st.form("form_editar", clear_on_submit=True):
        edit_id = st.number_input("ID del registro a editar", min_value=1, step=1)
        new_shipment_edit = st.text_input("Nuevo Shipment")
        new_referencia_edit = st.text_input("Nueva Referencia")
        new_marca_edit = st.text_input("Nueva Marca")
        # ... añade los campos que quieras editar ...
        submitted_edit = st.form_submit_button("Guardar cambios")

        if submitted_edit:
            with get_session() as session:
                reg = session.get(OCMaui, edit_id)
                if reg:
                    if new_shipment_edit:
                        reg.shipment = new_shipment_edit
                    if new_referencia_edit:
                        reg.referencia = new_referencia_edit
                    if new_marca_edit:
                        reg.marca = new_marca_edit
                    # ... más campos ...
                    try:
                        session.add(reg)
                        session.commit()
                        st.success("¡Registro editado con éxito!")
                        st.experimental_rerun()
                    except Exception as e:
                        st.error(f"Error al actualizar: {e}")
                else:
                    st.warning("No se encontró un registro con ese ID.")

    # 4) ELIMINAR (DELETE)
    st.subheader("Eliminar registro")
    del_id = st.text_input("ID a eliminar")
    if st.button("Eliminar"):
        if del_id:
            try:
                del_id_int = int(del_id)
                with get_session() as session:
                    reg = session.get(OCMaui, del_id_int)
                    if reg:
                        session.delete(reg)
                        session.commit()
                        st.success(f"Registro con ID {del_id_int} eliminado.")
                        st.experimental_rerun()
                    else:
                        st.warning("No se encontró un registro con ese ID.")
            except ValueError:
                st.error("El ID debe ser un número entero.")
        else:
            st.warning("Ingresa un ID para eliminar.")

def page_realizar_analisis():
    st.markdown("## Análisis de Subfamilias con Clasificación")

    # 1) Seleccionar año a analizar. Por defecto año actual (ajusta a tu preferencia).
    anio_predeterminado = datetime.now().year
    anio_seleccionado = st.number_input(
        "Seleccione año a analizar",
        min_value=1990,
        max_value=2100,
        value=anio_predeterminado,
        step=1
    )

    # 2) Subir archivo histórico (XLSB/XLSX)
    uploaded_file = st.file_uploader("Subir archivo histórico (.xlsb o .xlsx)", type=["xlsb", "xlsx"])

    if not uploaded_file:
        st.info("Por favor, sube el archivo histórico para continuar.")
        return

    # 3) Si ya tenemos los datos procesados en session_state, no volvemos a procesar
    if 'df_agrupado' in st.session_state:
        df_agrupado = st.session_state['df_agrupado']
        st.write("### Datos cargados previamente")
        st.dataframe(df_agrupado)
    else:
        # Botón para procesar
        if st.button("Procesar Análisis"):
            try:
                content = uploaded_file.read()
                import io
                if uploaded_file.name.endswith(".xlsb"):
                    df = pd.read_excel(io.BytesIO(content), engine='pyxlsb')
                else:
                    df = pd.read_excel(io.BytesIO(content))

                # Verificar columna 'fecha_despacho'
                if 'fecha_despacho' not in df.columns:
                    st.error("No se encontró la columna 'fecha_despacho' en el archivo.")
                    return
                
                # Mostrar los datos cargados antes de procesar las fechas
                st.write("### Datos cargados:")
                st.dataframe(df.head())  # Mostrar las primeras filas del dataframe

                # Convertir la fecha, asumiendo formato dd/mm/yyyy => dayfirst=True
                df['fecha_despacho'] = df['fecha_despacho'].apply(parsear_fecha)
                
                # Mostrar cuántas fechas se convirtieron
                n_valido = df['fecha_despacho'].notna().sum()
                n_total = len(df)
                st.write(f"Se convirtieron {n_valido} fechas correctamente de un total de {n_total} registros.")

                # Crear Mes y Año
                df['Mes'] = df['fecha_despacho'].dt.month
                df['Año'] = df['fecha_despacho'].dt.year

                # Filtrar por el año seleccionado
                df_anio = df[df['Año'] == anio_seleccionado].copy()
                cantidad_en_anio = len(df_anio)

                st.write(f"Registros filtrados para año {anio_seleccionado}: {cantidad_en_anio}")

                if df_anio.empty:
                    st.warning(f"No hay datos para el año {anio_seleccionado}. Verifica el formato de fecha.")
                    return

                # Asegurar que existe la columna de unidades y la columna 'Descripción Padre'
                if 'Cant_Unidad' not in df_anio.columns:
                    st.error("No se encontró la columna 'Cant_Unidad'.")
                    return
                if 'Descripción Padre' not in df_anio.columns:
                    st.error("No se encontró la columna 'Descripción Padre'.")
                    return

                # Extraer Sub Familia a partir de 'Descripción Padre'
                df_anio["Sub Familia"] = df_anio["Descripción Padre"].apply(extraer_descripcion)

                # Agrupar por Mes, Sub Familia
                agrupado = df_anio.groupby(['Mes', 'Sub Familia'], as_index=False)['Cant_Unidad'].sum()

                # Calcular total del mes y porcentaje
                agrupado['Total_Mes'] = agrupado.groupby('Mes')['Cant_Unidad'].transform('sum')
                agrupado['Porcentaje'] = ((agrupado['Cant_Unidad'] / agrupado['Total_Mes']) * 100).round(2)
                agrupado.sort_values(['Mes', 'Sub Familia'], inplace=True)

                # Crear la columna '&&' si se desea
                agrupado["&&"] = agrupado["Mes"].astype(str) + agrupado["Sub Familia"].astype(str)

                # Clasificar (Mezzanine o Selectivo) y comentarios
                def clasificar(row):
                    if row["Cant_Unidad"] > 450:
                        return "Mezzanine"
                    else:
                        return "Selectivo"

                agrupado["Clasificacion"] = agrupado.apply(clasificar, axis=1)

                def comentarios(row):
                    if row["Clasificacion"] == "Mezzanine":
                        return f"Subfamilia {row['Sub Familia']} con despachos significativos."
                    else:
                        return f"Subfamilia {row['Sub Familia']} con bajos despachos."

                agrupado["Comentarios"] = agrupado.apply(comentarios, axis=1)

                # Guardar el dataframe procesado en session_state
                st.session_state["df_agrupado"] = agrupado

                # Mostrar tabla
                st.markdown(f"### Resultado para Año {anio_seleccionado}")
                st.dataframe(agrupado.head(50))  # Vista previa

                # Mostrar algunos gráficos relevantes (ejemplo Altair)
                st.markdown("### Gráficos Relevantes")
                chart_data = agrupado.groupby("Sub Familia", as_index=False)["Cant_Unidad"].sum()
                chart = alt.Chart(chart_data).mark_bar().encode(
                    x=alt.X("Cant_Unidad:Q", title="Unidades Despachadas"),
                    y=alt.Y("Sub Familia:N", sort="-x"),
                    tooltip=["Sub Familia", "Cant_Unidad"]
                ).properties(width=600, height=400)
                st.altair_chart(chart, use_container_width=True)

                # Botón para exportar a Excel
                if st.button("Exportar Análisis a Excel"):
                    from io import BytesIO
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        agrupado.to_excel(writer, index=False, sheet_name="Análisis_Subfamilias")
                    output.seek(0)
                    st.download_button(
                        label="Descargar Excel",
                        data=output,
                        file_name=f"analisis_subfamilias_{anio_seleccionado}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            except Exception as e:
                st.error(f"Ocurrió un error procesando el archivo: {e}")

def page_consolidar_oc():
    icon = "📝"
    st.markdown(f"## {icon} Registro de OC´s")

    # ---------- BLOQUE DE 3 COLUMNAS PARA CARGA -----------    
    st.markdown("### Cargar Documentos")
    col_consolidado, col_curva, col_plantilla = st.columns(3)

    # Col 1: CSV Consolidado (con métrica de cantidad)
    with col_consolidado:
        st.markdown("**Cargar Ordenes de Compra (CSV)**")
        uploaded_files = st.file_uploader(
            "Subir uno o más CSV (Consolidado)",
            type=["csv"],
            accept_multiple_files=True,
            key="csv_consolidado",
            label_visibility="collapsed"  # Ocultar label por defecto
        )
        n_consolidado = len(uploaded_files) if uploaded_files else 0
        st.write(f"**Documentos**: {n_consolidado} archivos cargados")

    # Col 2: CSV Curva Artículo (sin métrica)
    with col_curva:
        st.markdown("**Cargar Curva Artículo**")
        curva_articulo_file = st.file_uploader(
            "Cargar Curva Artículo",
            type=["csv"],
            key="csv_curva",
            label_visibility="collapsed"
        )

    # Col 3: XLSM Plantilla Explosión (sin métrica)
    with col_plantilla:
        st.markdown("**Plantilla Explosión (XLSM)**")
        plantilla_explosion_file = st.file_uploader(
            "Cargar Plantilla Explosión Maui (XLSM)",
            type=["xlsm"],
            key="xlsm_plantilla",
            label_visibility="collapsed"
        )
        if plantilla_explosion_file:
            st.session_state["plantilla_explosion_file"] = plantilla_explosion_file

    # ---------- BLOQUE DE 3 COLUMNAS PARA DATOS CONTENEDOR/REFERENCIA/FECHA -----------    
    st.markdown("### Datos de Contenedor")
    col_contenedor, col_referencia, col_fecha = st.columns(3)

    with col_contenedor:
        contenedor = st.text_input("Contenedor:")

    with col_referencia:
        referencia = st.text_input("Referencia:")

    with col_fecha:
        fecha_recepcion = st.date_input("Fecha de Recepción:", datetime.now())

    # ---------- VALIDACIÓN PARA PROCESAR ----------
    if uploaded_files and curva_articulo_file:
        # 1. Procesar CSV(s) Consolidados
        lista_df = []
        for upf in uploaded_files:
            try:
                df_temp = pd.read_csv(upf)
                lista_df.append(df_temp)
            except Exception as e:
                st.error(f"Error al leer {upf.name}: {e}")

        if lista_df:
            df_consolidado = pd.concat(lista_df, ignore_index=True)
            df_consolidado.insert(0, "Shipment", contenedor)
            df_consolidado.insert(1, "Referencia", referencia)
            df_consolidado.insert(2, "Fecha de Recepción", fecha_recepcion if fecha_recepcion else "")

            st.success("Archivos consolidados procesados correctamente.")

            # Buscar columna 'Descripcion' o 'Descripción'
            desc_cols = [c for c in df_consolidado.columns if c.lower() in ["descripcion", "descripción"]]
            if not desc_cols:
                st.error("No se encontró la columna 'Descripcion' en los CSV.")
                return
            desc_col = desc_cols[0]

            # Funciones auxiliares
            df_consolidado["Subfamilias"] = df_consolidado[desc_col].apply(extraer_descripcion)
            df_consolidado["Código Marca"] = df_consolidado.apply(
                lambda row: extraer_codigo_marca(row[desc_col], row["Subfamilias"]),
                axis=1
            )
            df_consolidado["Marca"] = df_consolidado["Código Marca"].apply(calcular_marca)
            df_consolidado["Zona"] = df_consolidado["Marca"].apply(calcular_zona)
            df_consolidado["Tipo_Pack"] = df_consolidado[desc_col].apply(calcular_tipo_pack)

            # 2. Procesar CSV Curva Artículo
            try:
                df_curva_articulo = pd.read_csv(curva_articulo_file)
                df_curva_articulo = calcular_factor_por_caja(df_curva_articulo)

                df_consolidado = calcular_factor_caja(df_consolidado, df_curva_articulo)
                df_consolidado = calcular_qty_inners(df_consolidado)
                df_consolidado = calcular_qty_unidades(df_consolidado)

                st.success("Curva Artículo procesada correctamente.")

                # Generar DataFrames derivados
                df_f_recepcion = generar_df_f_recepcion(df_consolidado)
                df_f_expl_unid = generar_df_f_expl_unid(df_consolidado)
                df_expl_inner  = generar_df_expl_inner(df_f_recepcion, df_curva_articulo)

                st.session_state["df_consolidado"]    = df_consolidado
                st.session_state["df_curva_articulo"] = df_curva_articulo
                st.session_state["df_f_recepción"]    = df_f_recepcion
                st.session_state["df_f_expl_unid"]    = df_f_expl_unid
                st.session_state["df_expl_inner"]     = df_expl_inner

                # Inicializamos
                if "contenedor_registrado" not in st.session_state:
                    st.session_state["contenedor_registrado"] = False
                if "show_toast" not in st.session_state:
                    st.session_state["show_toast"] = False

                # Título + Botón
                col_title, col_reg = st.columns([0.7, 0.3])
                with col_title:
                    st.markdown("### Tabla Consolidado OC's (Final)")
                with col_reg:
                    if not st.session_state["contenedor_registrado"]:
                        if st.button("Registrar Contenedor"):
                            # ========================
                            # MAPEO DE COLUMNAS (Paso 3)
                            # ========================
                            column_map = {
                                "Shipment": "shipment",
                                "Referencia": "referencia",
                                "Fecha de Recepción": "fecha_recepcion",
                                "Cliente": "cliente",
                                "Proveedor": "proveedor",
                                "Direccion": "direccion",
                                "No Factura": "nro_factura",
                                "Fecha Limite": "fecha_limite",
                                "Fecha Factura": "fecha_factura",
                                "Familia De Producto": "familia_producto",
                                "Num Producto": "num_producto",
                                "Descripcion": "descripcion",
                                "Producto Nuevo": "producto_nuevo",
                                "Huella": "huella",
                                "Huella Default": "huella_default",
                                "Recibo Habilitado": "recibo_habilitado",
                                "Cantidad Esperada": "cantidad_esperada",
                                "Identificada": "identificada",
                                "Cant Cajas": "cant_cajas",
                                "Saldos Un": "saldos_un",
                                "Vol M3": "vol_m3",
                                "Articulo Padre": "articulo_padre",
                                "Recibida": "recibida",
                                "Subfamilias": "subfamilia",
                                "Código Marca": "codigo_marca",
                                "Marca": "marca",
                                "Zona": "zona",
                                "Tipo_Pack": "tipo_pack",
                                "Factor_Caja": "factor_caja",
                                "Qty_Inners": "qty_inner",
                                "Qty_Unidades": "qty_unidades"
                            }

                            # Creamos df_renamed
                            df_renamed = df_consolidado.rename(columns=column_map)

                            # (Llamar a tu función que guarda en BD, por ejemplo):
                            try:
                                guardar_contenedor_bd(df_renamed)
                                st.success("¡Datos guardados en BD correctamente!")
                            except Exception as err:
                                st.error(f"Error al guardar en BD: {err}")

                            st.session_state["contenedor_registrado"] = True
                            st.session_state["show_toast"] = True
                    else:
                        st.success("Contenedor ya registrado")

                # Toast (opcional)
                if st.session_state["show_toast"]:
                    toast_html = """
                    <style>
                    .toast-container {
                        position: fixed;
                        bottom: 20px;
                        right: 20px;
                        background-color: #fff;
                        padding: 20px;
                        border-radius: 8px;
                        box-shadow: 0 2px 6px rgba(0,0,0,0.2);
                        z-index: 9999;
                        display: flex;
                        align-items: center;
                    }
                    .toast-container img {
                        margin-right: 10px;
                    }
                    </style>
                    <div class="toast-container" id="myToast">
                      <img src="https://cdn-icons-png.flaticon.com/512/190/190411.png" width="40"/>
                      <div>
                        <strong>Datos de Contenedor registrados correctamente</strong>
                      </div>
                    </div>
                    <script>
                    setTimeout(function(){
                       var t = document.getElementById("myToast");
                       if(t){ t.style.display = "none"; }
                    }, 2000);
                    </script>
                    """
                    st.markdown(toast_html, unsafe_allow_html=True)
                    st.session_state["show_toast"] = False

                # --- Resumen Didáctico ---
                if all(col in df_consolidado.columns for col in ["Qty_Inners", "Qty_Unidades"]):
                    total_inners = df_consolidado["Qty_Inners"].sum()
                    total_unidades = df_consolidado["Qty_Unidades"].sum()

                    # Sección de Resumen
                    st.markdown("## Resumen de Inners y Unidades")
                    st.caption("Vista rápida de totales y desglose por marca / zona")

                    # Tarjetas horizontales
                    c_inners, c_units = st.columns(2)
                    with c_inners:
                        st.markdown(f"""
                        <div style="background-color:#f8f9fa; padding:15px; border-radius:8px;">
                          <h4 style="margin:0;">Total Inners</h4>
                          <p style="font-size:35px; margin:0; color:#007BFF;">{int(total_inners):,}</p>
                        </div>
                        """, unsafe_allow_html=True)
                    with c_units:
                        st.markdown(f"""
                        <div style="background-color:#f8f9fa; padding:15px; border-radius:8px;">
                          <h4 style="margin:0;">Total Unidades</h4>
                          <p style="font-size:35px; margin:0; color:#28a745;">{int(total_unidades):,}</p>
                        </div>
                        """, unsafe_allow_html=True)

                    # Tabla por Marca y Zona con un estilo
                    st.markdown("#### Desglose por Marca y Zona")
                    df_marca_zona = (
                        df_consolidado
                        .groupby(["Marca", "Zona"], as_index=False)[["Qty_Inners", "Qty_Unidades"]]
                        .sum()
                    )
                    df_marca_zona["Qty_Inners"] = df_marca_zona["Qty_Inners"].astype(int)
                    df_marca_zona["Qty_Unidades"] = df_marca_zona["Qty_Unidades"].astype(int)

                    st.dataframe(
                        df_marca_zona.style.highlight_max(subset=["Qty_Inners", "Qty_Unidades"], color="#d4edda")
                    )

                    if "Subfamilias" in df_consolidado.columns:
                        st.markdown("#### Gráfico de Subfamilias (Inners vs Unidades)")
                        df_subfam = (
                            df_consolidado
                            .groupby("Subfamilias", as_index=False)[["Qty_Inners", "Qty_Unidades"]]
                            .sum()
                        )
                        df_subfam["Total"] = df_subfam["Qty_Inners"] + df_subfam["Qty_Unidades"]
                        df_subfam.sort_values("Total", ascending=False, inplace=True)

                        # Gráfico con Altair
                        chart_data = df_subfam.melt(
                            id_vars="Subfamilias",
                            value_vars=["Qty_Inners", "Qty_Unidades"],
                            var_name="Tipo",
                            value_name="Cantidad"
                        )
                        chart = alt.Chart(chart_data).mark_bar().encode(
                            x=alt.X("Cantidad:Q", title="Cantidad"),
                            y=alt.Y("Subfamilias:N", sort="-x"),
                            color="Tipo:N",
                            tooltip=["Subfamilias", "Tipo", "Cantidad"]
                        ).properties(width=600, height=400)

                        st.altair_chart(chart, use_container_width=True)
                    else:
                        st.warning("No existe la columna 'Subfamilias' para el resumen.")
                else:
                    st.warning("No se encontraron columnas 'Qty_Inners' y/o 'Qty_Unidades' para el resumen.")

                # ---- Mostrar la tabla principal
                st.markdown("### Tabla Consolidado (Detalle)")
                mostrar_y_descargar_dataframe(df_consolidado, "consolidado_oc_final")

                # Sección "Consultar Formatos Generados" y "Exportar"
                if st.session_state["contenedor_registrado"]:
                    st.markdown("### Consultar Formatos Generados")
                    opciones = [
                        "df_f_expl_unid",
                        "df_curva_articulo",
                        "df_f_recepción",
                        "df_expl_inner"
                    ]
                    seleccion = st.selectbox("Selecciona un DataFrame para visualizar:", opciones)

                    if seleccion in st.session_state:
                        df_seleccionado = st.session_state[seleccion]
                        st.info(f"Mostrando: {seleccion}")
                        mostrar_y_descargar_dataframe(df_seleccionado, seleccion)
                    else:
                        st.warning("Aún no se ha generado el DataFrame seleccionado.")

                    if st.button("Exportar Plantilla Explosión"):
                        try:
                            if "plantilla_explosion_file" not in st.session_state:
                                st.error("No se encontró la plantilla XLSM en session_state.")
                                return

                            df_f_recep = st.session_state["df_f_recepción"]
                            df_f_unid  = st.session_state["df_f_expl_unid"]
                            df_inner   = st.session_state["df_expl_inner"]

                            contenedor_val = df_consolidado["Shipment"].unique()[0]
                            referencia_val = df_consolidado["Referencia"].unique()[0]
                            fecha_val      = df_consolidado["Fecha de Recepción"].unique()[0]

                            if isinstance(fecha_val, pd.Timestamp):
                                fecha_str = fecha_val.strftime("%Y%m%d")
                            else:
                                fecha_str = str(fecha_val)

                            plantilla_bytes = st.session_state["plantilla_explosion_file"]
                            in_memory_file = BytesIO(plantilla_bytes.getvalue())
                            wb = load_workbook(in_memory_file, keep_vba=True)

                            sheet_unid = wb["df_f_expl_unid"]

                            # Estos valores en las celdas I2, J2, K2 permanecen igual:
                            sheet_unid["I2"] = contenedor_val
                            sheet_unid["J2"] = referencia_val
                            sheet_unid["K2"] = str(fecha_val)

                            # Definimos la celda de inicio para pegar los datos (C9 => fila 9, columna 3)
                            start_row = 11
                            start_col = 3

                            # 1. Convertir el DataFrame en filas (sin encabezados, sin índice)
                            rows = dataframe_to_rows(df_f_unid, index=False, header=False)

                            # 2. Pegar en bloque en la hoja:
                            for r_idx, row_data in enumerate(rows, start=start_row):
                                for c_idx, cell_value in enumerate(row_data, start=start_col):
                                    sheet_unid.cell(row=r_idx, column=c_idx, value=cell_value)

                            # Eliminar filas vacías del DataFrame df_f_recep antes de pegarlas
                            df_f_recep_clean = df_f_recep.dropna(how='all')  # Elimina filas completamente vacías

                            # Escribir todos los datos en bloque (a partir de la fila 11)
                            sheet_recep = wb["df_f_recepción"]
                            start_row_recep = 11  # Comenzamos en la fila 11 de la columna C para df_f_recep

                            # Convertir el DataFrame limpio a filas de Excel y escribirlas en bloque
                            for r_idx, row in enumerate(dataframe_to_rows(df_f_recep_clean, index=False, header=False), start=start_row_recep):
                                for c_idx, value in enumerate(row, start=3):  # Comienza a pegar desde la columna 3 (columna C)
                                    sheet_recep.cell(row=r_idx, column=c_idx, value=value)

                            start_row1 = 11
                            sheet_inner = wb["df_expl_inner"]
                            # Pegar los datos en bloque desde la celda C11
                            for i, row_data in df_inner.iterrows():
                                for j, value in enumerate(row_data):
                                    sheet_inner.cell(row=start_row1 + i, column=start_col + j, value=value)

                            out_file = BytesIO()
                            file_name = f"Explosión_Maui_{contenedor_val}_{referencia_val}_{fecha_str}.xlsm"
                            wb.save(out_file)
                            out_file.seek(0)

                            st.download_button(
                                label="Descargar Archivo Explosión",
                                data=out_file,
                                file_name=file_name,
                                mime="application/vnd.ms-excel.sheet.macroEnabled.12"
                            )
                            st.success(f"Plantilla exportada correctamente: {file_name}")
                                 
                        except Exception as e:
                            st.error(f"Error al exportar la plantilla: {e}")

                else:
                    st.warning("Por favor, registre el contenedor para habilitar la exportación.") 

            except Exception as e:
                st.error(f"Ocurrió un error al cargar el archivo de Curva Artículo: {e}")
        else:
            st.warning("No se pudo consolidar ningún archivo.")
    else:
        st.warning("Por favor, sube los archivos CSV y el archivo de Curva Artículo para continuar.")



###############################################################################
# 7. FUNCIÓN PRINCIPAL (NAVEGACIÓN)
###############################################################################
# Función principal
def main():
    # (1) Todo tu bloque de autenticación (OneDrive) se mantiene igual:
    auth_code = get_auth_code_from_url()
    if auth_code:
        # Intercambiamos el código por un token de acceso
        result = get_access_token_from_code(auth_code)
        if "access_token" in result:
            access_token = result["access_token"]
            st.success("Autenticación exitosa. Accediendo a OneDrive...")

            # Mostrar que está conectado a OneDrive
            email = get_user_email(access_token)
            if email:
                st.write(f"Conectado con OneDrive como: {email}")
            else:
                st.error("No se pudo obtener el correo electrónico del usuario.")

            # Listar archivos de OneDrive
            files = list_onedrive_files(access_token)
            if files:
                st.write("Archivos en OneDrive:")
                for file in files:
                    st.write(file["name"])
            else:
                st.error("No se encontraron archivos en OneDrive.")
        else:
            st.error("Error al obtener el token de acceso.")
    else:
        # Si no hay código, mostrar el botón de inicio de sesión
        st.write("Para acceder a los archivos de OneDrive, por favor inicie sesión.")
        login_button()

    # (2) Llamamos a la función que crea el menú, sin manipular manualmente el session_state:
    opcion = radio_menu_con_iconos()

    # (3) El resto de tu aplicación con if/elif en base a 'opcion':
    with st.container():
        st.markdown("<div class='main-container'>", unsafe_allow_html=True)

        if opcion == "Inicio":
            page_home()
        elif opcion == "Realizar Análisis":
            page_realizar_analisis()
        elif opcion == "Registro de OC´s":
            page_consolidar_oc()
        elif opcion == "Consultar BD":
            page_consultar_bd()
        elif opcion == "Salir":
            # Ojo: si usas 'MENU_OPCIONES["Sadeflir"]' pero tu dict se llama "Salir",
            # esto generará error. Asegúrate de que coincida la clave. 
            # Por ejemplo, si tu menú usa 'Salir', hazlo así:
            icon = MENU_OPCIONES["Salir"]
            st.markdown(f"## {icon} Salir")
            st.warning("Has salido del Sistema de Gestión de Abastecimiento. "
                       "Cierra la pestaña o selecciona otra opción.")

        st.markdown("</div>", unsafe_allow_html=True)

if __name__ == "__main__":
    main()
